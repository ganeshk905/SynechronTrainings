import sqlite3
con=sqlite3.connect('mydb')
#con=otherDB.connect(úser='', password='', host='', port='', database='') #import for othe DB flavours...ORacle, sQL,
cur=con.cursor()
query='create table if not exists log_data(ip VARCHAR(100), dt VARCHAR(100), pics VARCHAR(100), wb VARCHAR(100))'
cur.execute(query)

import re
F=open(r'C:\training\log\out4.txt')
for line in F:
    m=re.match('(\d{3}\.\d{1,3}\.[0-9]{3}\.\d{1,3}).*(\d{2}/[a-zA-Z]{3}/\d{4}).*"GET\s+/(pics/\w+\.\w+)?.*http://(\w+\.\w+\.\w+)/\w+', line)#When you put the pattern in (), it becomes a group..you can have multiple groups and refer them through index, here in this example we have created only 1 group, hence referred through group(1)
    if m!=None:
        ip=m.group(1)
        dt=m.group(2)
        pc=m.group(3)
        wb=m.group(4)
        if pc!= None:
         pc=pc.lstrip('pics/')
        else:
         pc='No Image'
        print(ip)
        print(dt)
        print(pc)
        print(wb)
        query="insert into log_data values('{}','{}','{}','{}')".format(ip,dt,pc,wb)
        cur.execute(query)

con.commit()
query='select * from log_data'
cur.execute(query)
result=cur.fetchall()
print('Result =', result)


#put the DB records into excel or csv using modules, earlier we have used print optionnnnn

import csv
F=open(r'C:\training\log\out7.csv', 'w', newline='')
wtr=csv.writer(F)
for i in result:
    wtr.writerow(i)
F.close()

F=open(r'C:\training\log\out7.csv', newline='')
rdr=csv.reader(F)
rdr=list(rdr) #to use it second time
for i in rdr:
    print('i =', i)
F.close()

import openpyxl#module for excel
wb=openpyxl.Workbook()
#wb.save(r'C:\training\log\out7.xlsx') Just to see file is getting created, we have written it here, we will use this at the end.

ws1=wb.active
#ws1=wb['Sheet1']
ws2=wb.create_sheet('Report')
#for column specification
ws1['A1']=100
ws2['A1']=100
#for row and column specification
ws1.cell(row=2, column=1).value=200
ws2.cell(row=2, column=1).value=200
#for simply append
L=[10,20,30]
ws1.append(L)
ws2.append(L)

for r in rdr:
    ws1.append(r)
    ws2.append(r)
wb.save(r'C:\training\log\out7.xlsx')

#To read from excel and publish, first open workbook
wb=openpyxl.load_workbook(r'C:\training\log\out7.xlsx')
sheets=wb.sheetnames
ws=wb['Report']
tr=ws.max_row#total number of used row
tc=ws.max_column#total number of used columns#
#print(ws)

ws2=wb.create_sheet('Report2')

for r in range(3, tr+1): #starting from 3rd row to total used rows
    for c in range(2, tc+1): #starting from 2nd column to total used column
        ws2.cell(row=r-2, column=c-1).value=ws.cell(row=r, column=c).value
wb.save(r'C:\training\log\out7.xlsx')

for i in ws['A']:
    print(i.value)
for j in ws['2']:
    print(j.value)
for k in ws['A:C']:
    for l in k:
        print('l = ', l.value)

wb.close()
con.close()

wb=openpyxl.load_workbook(r'C:\training\log\out8.xlsx')
ws1=wb.active
ws2=wb.create_sheet('Area')
ws3=wb.create_sheet('Bar')

from openpyxl.chart import AreaChart, BarChart, Reference

chart1=AreaChart()
chart2=BarChart()
cat=Reference(ws1, min_row=1, max_row=10, min_col=1) #first column, row 1 to 10 would be category
data=Reference(ws1, min_row=1, max_row=10, min_col=2, max_col=3)#data reference , you will have 2 things - category and data
chart1.set_categories(cat)
chart2.set_categories(cat)
chart1.add_data(data)
chart2.add_data(data)

ws2.add_chart(chart1)
ws3.add_chart(chart2)

wb.save(r'C:\training\log\out8.xlsx')





'''
#ws['A1']....1 cell
#ws['A'].....1 column
#ws['A:C']....A to c columns
#ws['2']....1 row
#ws['2:10']...2 to 10 rows
'''

#\d for digit but 1 digit, if you want 3, put it as \d{3}...next ..if you ant to match any single character, use '.' but if you want tro match '.' pattern, use \.

''''
1) Also another way is use [0-5]{3}, it means 3 digits between range 0 to 5
2) \d* 0 or more digits
3) \D  non-digit
4) .  Any single character
5) \d+ 1 or more digits
6) \s+ 1 or more space
7) .* 0 or more characters
8) \w [0-9a-zA-z_] it has 0-9 plus a-z plus Ato Z plus '_' [word char]
9) \W [^a-zA-Z0-9_] it does not have 0-9 plus a-z plus Ato Z plus '_'[non-word char]
10) \w+ 
11) + 1 or more
12) * 0 or more
13) ? 0 or 1
14) ^ beginning
15) $ ending
16) \d? 0 or 1 digits
'''