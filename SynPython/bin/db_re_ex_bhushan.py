#DB - for every db diff module need to import

#module for sqlite3
import sqlite3

#connection details - for sqlite, if db not there, it will create the db
con=sqlite3.connect('mydb')
#for other sql db etc, in live project below will be conn string
#con=otherDB.connect(user='',password='',host='',port='',database='')

cur=con.cursor()
query='create table if not exists log_data(ip VARCHAR(100), dt VARCHAR(100), pics VARCHAR(100), wb VARCHAR(100))'
cur.execute(query)

#module for regex
import re

'''
\d => digit
\D => non-digit
\. => Any char
\w => word char [a-zA-Z0-9_] (underscore and alpha numeric)
\W => non word  [^a-zA-Z0-9_] (apart from underscore and alpha numeric)
\s => space
\S => Non space
[] => mention a-z, A-Z, 0-9, etc
{} => mention how many times it is occuring
+ => 1 or more
* => 0 or more
? => 0 or 1
^ => begining (if ^ inside [^] then denotes not eg: [^a-zA-Z0-9_] (apart from underscore and alpha numeric))
$ => ending
'''

'''\d => digit, \d{3}=>3 digits, .=> any single char, \.=> match exact . , \d{1,3} => 1 to 3 digits (min 1 n max 3 digits),
    [0-9]{3} => 3 digits bet 0 to 9(can mention 0-5), * => 0 or more occurance, + => 1 or more, ? => 0 or 1 (\d*, \d+, \d? etc)
    [a-zA-z]{3} => 3 chars, / => hardcoded /, \s => space, \s+ => one or more space, \w => [0-9a-zA-Z_] (underscore and alpha numeric)
    [a-zA-z0-9.]+ => 1 or more comb of (a-zA-z0-9.)

    (\d{3}\.\d{1,3}\.[0-9]{3}\.\d{3}) => 123.123.123.123
    (\d{2}/[a-zA-z]{3}/\d{4}) => 20/Jun/2018
    "GET\s+/(pics/\w+\.\w+)? => "GET /pics/5star2000.gif or "GET /asctortf/ HTTP/1.0 => (note: ? => 0 or 1, may or may not, if not found then returns None)
    http://([a-zA-z0-9.]+) =>  www.jafsoft.com 
    '''

'''
F=open(r'C:\Bhushan\Python\log\plog.txt')
for line in F:
    m = re.match('(\d{3}\.\d{1,3}\.[0-9]{3}\.\d{3}).*(\d{2}/[a-zA-z]{3}/\d{4}).*"GET\s+/(pics/\w+\.\w+)?.*http://([a-zA-z0-9.]+)', line)

    if m != None:
        ip = m.group(1)
        #print(ip)
        dt=m.group(2)
        #print(dt)
        pc = m.group(3)
        if pc != None:
            pc=pc.lstrip('pics/')
        else:
            pc='No Image'
        #print(pc)
        wb = m.group(4)
        #print(wb)
        query = "insert into log_data values('{}','{}','{}','{}')".format(ip,dt,pc,wb)
        cur.execute(query)

con.commit() #outside for
'''

#run below after commiting con.commit else it will add records repeatedly
query='select * from log_data'
cur.execute(query)
result = cur.fetchall() #list of tuple
print(result)

#module for csv
import csv

F=open(r'C:\Bhushan\Python\log\out6.csv', 'w', newline='')
wtr = csv.writer(F)

for i in result:
    wtr.writerow(i)
F.close()

F=open(r'C:\Bhushan\Python\log\out6.csv', newline='')
rdr = csv.reader(F) #returns generator object...2nd access will give empty
rdr=list(rdr) #do not convert to list...here we want that data in next example of excel...hence saving in list
for i in rdr:
    print('i= ', i)
F.close()

#openpyxl and pandas used for excel operations
#Use pandas most of time...also preferred for huge no records , openpyxl => reding/write/beautifn
import openpyxl
wb = openpyxl.Workbook()
ws1 = wb.active #will give active sheet
#ws1=wb['Sheet']
ws2=wb.create_sheet('Report') #will create n give Report sheet

#1ST WAY WRITE
ws1['A1']=100 #write to w1 Sheet
ws2['A1']=100 #write to w2 Report Sheet

#2nd WAY WRITE
ws1.cell(row=2, column=1).value = 200
ws2.cell(row=2, column=1).value = 200

#3rd WAY WRITE
L=[10,20,30]
ws1.append(L) #will write at last unused row in diff cells
ws2.append(L)

for r in rdr:
    ws1.append(r)
    ws2.append(r)

wb.save(r'C:\Bhushan\Python\log\out7.xlsx')

#to load/open excel file
wb=openpyxl.load_workbook(r'C:\Bhushan\Python\log\out7.xlsx')
#to get sheets from workbook
sheets = wb.sheetnames
ws=wb['Report']
tr=ws.max_row #row count
tc=ws.max_column #column count
print(tr,tc)

#customized : to copy certain row n values from one sheet and writing to another sheet
ws2=wb.create_sheet('Report2')
for r in range(3, tr+1):
    for c in range(2, tc+1):
        ws2.cell(row=r-2,column=c-1).value = ws.cell(row=r, column=c).value
wb.save(r'C:\Bhushan\Python\log\out7.xlsx')

#print values in A column
for i in ws['A']:
    print(i.value)

#print values in 2nd row
for j in ws['2']:
    print(j.value)

print('================================================')
#print values in 2nd row
for k in ws['A:C']: #column values in tuple
    for l in k:
        print(l.value)

wb.close()
con.close()
#ws['A1'] => 1 cell
#ws['A'] => 1 column
#ws['A:C'] => A to C columns
#ws['2'] => 1 row
#ws['2:10'] => 2 to 10 rows

wb=openpyxl.load_workbook(r'C:\Bhushan\Python\log\chart.xlsx')
ws1=wb['Sheet1']
ws2=wb.create_sheet('Area')
ws3=wb.create_sheet('Bar')

from openpyxl.chart import AreaChart,BarChart,Reference

#Chart Obj => Category: x-axis, data: y-axis

#expore other charts by urself
chart1=AreaChart()
chart2=BarChart()

cat=Reference(ws1, min_row=1, max_row=10,min_col=1)
data=Reference(ws1,min_row=1,max_row=10,min_col=2,max_col=3)

chart1.set_categories(cat)
chart1.add_data(data)

chart2.set_categories(cat)
chart2.add_data(data)

ws2.add_chart(chart1)
ws3.add_chart(chart2)
wb.save(r'C:\Bhushan\Python\log\chart.xlsx')




